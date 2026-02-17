import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

FILE_NAME = "Factory_ERP.xlsx"


def create_erp_file():

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")

    def add_sheet(title, headers):
        ws = wb.create_sheet(title)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
            ws.column_dimensions[chr(64 + col)].width = 22

    # Dashboard
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "FACTORY ERP SYSTEM"
    ws["A1"].font = Font(size=16, bold=True)

    # Masters
    add_sheet("Customer Master",
              ["Customer ID","Customer Name","Contact","Email","Address","GST","Payment Terms"])

    add_sheet("Vendor Master",
              ["Vendor ID","Vendor Name","Material Supplied","Phone","GST","Lead Time"])

    add_sheet("Item Master",
              ["Item Code","Item Name","Category","Unit","Standard Cost","Min Stock"])

    # BOM
    add_sheet("BOM",
              ["Product Code","Raw Material Code","Quantity Required","Unit"])

    # Purchase
    add_sheet("Purchase Orders",
              ["PO Number","Date","Vendor","Item","Quantity","Rate","Expected Delivery","Status"])

    add_sheet("GRN",
              ["GRN Number","Date","PO Number","Item","Received Qty","Accepted Qty","Rejected Qty"])

    # Production
    add_sheet("Production Plan",
              ["Plan Number","Date","Product","Planned Qty","Start Date","End Date","Status"])

    add_sheet("Job Cards",
              ["Job Card No","Plan Number","Product","Quantity","Machine","Operator","Start Time","End Time","Status"])

    add_sheet("Production Entry",
              ["Date","Job Card No","Product","Produced Qty","Rejected Qty","Shift"])

    # Inventory
    add_sheet("Raw Material Inventory",
              ["Item Code","Opening Stock","Received","Issued","Balance"])

    add_sheet("Finished Goods Inventory",
              ["Product","Opening Stock","Produced","Dispatched","Balance"])

    # Sales
    add_sheet("Sales Orders",
              ["SO Number","Customer","Product","Quantity","Delivery Date","Status"])

    add_sheet("Dispatch",
              ["Dispatch No","SO Number","Product","Quantity","Dispatch Date","Transport","Invoice No"])

    add_sheet("Invoices",
              ["Invoice No","Customer","Amount","Paid","Pending","Due Date"])

    wb.save(FILE_NAME)


def ensure_erp_exists():
    if not os.path.exists(FILE_NAME):
        create_erp_file()
